# -*- coding: utf-8 -*-
"""
딥파인_PR리포트/files/make_month_tab.py + classify_articles.py 의 핵심 로직을
웹앱(app.py)에서 재사용하기 좋게 함수로 뽑아놓은 모듈.
CLI 스크립트(argparse/print) 부분만 제거했고, 실제 동작 로직은 그대로 옮김.
"""
import copy
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
import json
from datetime import datetime, timedelta, timezone

from openpyxl.cell.cell import MergedCell

from media_order import media_sort_key

NAVER_URL = "https://openapi.naver.com/v1/search/news.json"
KST = timezone(timedelta(hours=9))

MEDIA_MAP = {
    "ajunews.com": "아주경제", "mt.co.kr": "머니투데이", "unicornfactory.co.kr": "유니콘팩토리",
    "ddaily.co.kr": "디지털데일리", "industrynews.co.kr": "인더스트리뉴스", "fajournal.com": "FA저널",
    "hellot.net": "헬로티", "klnews.co.kr": "물류신문", "epnc.co.kr": "테크월드",
    "datanet.co.kr": "데이터넷", "thelec.kr": "디일렉", "ezyeconomy.com": "이지경제",
    "hansbiz.co.kr": "한스경제", "cstimes.com": "컨슈머타임스", "public25.com": "퍼블릭타임스",
    "venturesquare.net": "벤처스퀘어", "newsworks.co.kr": "뉴스웍스", "ebn.co.kr": "EBN",
    "tech42.co.kr": "테크42", "newstap.co.kr": "뉴스탭", "monthlypeople.com": "월간인물",
    "denews.co.kr": "디지털경제뉴스", "ksg.co.kr": "코리아쉬핑가제트", "polinews.co.kr": "폴리뉴스",
    "smedaily.co.kr": "중소기업신문", "the-biz.co.kr": "THE Biz", "sentv.co.kr": "서울경제TV",
    "bosa.co.kr": "의학신문", "newsis.com": "뉴시스", "edaily.co.kr": "이데일리",
    "s-journal.co.kr": "S저널", "byline.network": "바이라인네트워크", "biz.chosun.com": "조선비즈",
    "zdnet.co.kr": "지디넷코리아", "aitimes.com": "AI타임스", "dealsite.co.kr": "딜사이트",
    "etnews.com": "전자신문", "thebell.co.kr": "더벨", "fashionbiz.co.kr": "패션비즈",
    "apparelnews.co.kr": "어패럴뉴스",
    "platum.kr": "플래텀",
    "bloter.net": "블로터",
    "technoa.co.kr": "테크노아",
}

# Streamlit Cloud는 파일시스템이 매 실행마다 초기화될 수 있어서 디스크 캐시 대신
# 프로세스 메모리 캐시만 사용 (같은 세션 내 반복 조회만 절약, 손해 없음)
_media_cache = {}


def clean_html(s):
    s = re.sub(r"<.*?>", "", s or "")
    return (
        s.replace("&quot;", '"').replace("&amp;", "&")
        .replace("&lt;", "<").replace("&gt;", ">").replace("&apos;", "'").strip()
    )


def _domain_of(url):
    m = re.search(r"https?://(?:www\.)?([^/]+)", url or "")
    return m.group(1) if m else ""


def _decode_html(raw: bytes) -> str:
    """국내 뉴스 사이트 중 EUC-KR/CP949 인코딩을 쓰는 곳이 있어서, UTF-8로 무작정
    디코딩하면 매체명이 깨짐(예: boannews.com -> 'ȴ'). 페이지에 선언된 charset을
    먼저 찾아보고, 없으면 흔한 인코딩들을 strict하게 시도해본 뒤 최후에만 ignore."""
    head = raw[:2000].decode("latin-1", errors="ignore")
    m = re.search(r'charset=["\']?\s*([\w-]+)', head, re.IGNORECASE)
    candidates = []
    if m:
        enc = m.group(1).lower()
        candidates.append({"euckr": "euc-kr", "ms949": "cp949"}.get(enc, enc))
    candidates += ["utf-8", "cp949", "euc-kr"]
    for enc in candidates:
        try:
            return raw.decode(enc, errors="strict")
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="ignore")


def _fetch_site_name(url):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; ReportBot/1.0)"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            raw = resp.read(200_000)
        html = _decode_html(raw)
        m = re.search(r'<meta[^>]+property=["\']og:site_name["\'][^>]+content=["\']([^"\']+)["\']', html, re.IGNORECASE)
        if not m:
            m = re.search(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:site_name["\']', html, re.IGNORECASE)
        if m:
            return clean_html(m.group(1))
    except Exception:
        pass
    return None


def media_from_link(link):
    domain = _domain_of(link)
    if not domain:
        return ""
    if domain in MEDIA_MAP:
        return MEDIA_MAP[domain]
    bare = re.sub(r"^[^.]+\.", "", domain)
    if bare in MEDIA_MAP:
        return MEDIA_MAP[bare]
    if domain in _media_cache:
        return _media_cache[domain]
    name = _fetch_site_name(link)
    if name:
        _media_cache[domain] = name
        return name
    m = re.search(r"([^./]+)\.", domain)
    return m.group(1) if m else domain


def fetch_all_naver(keyword, client_id, client_secret):
    results = []
    start = 1
    while start <= 1000:
        params = urllib.parse.urlencode({"query": keyword, "display": 100, "start": start, "sort": "date"})
        req = urllib.request.Request(f"{NAVER_URL}?{params}")
        req.add_header("X-Naver-Client-Id", client_id)
        req.add_header("X-Naver-Client-Secret", client_secret)
        try:
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"네이버 API 호출 실패: HTTP {e.code}: {e.read().decode('utf-8', errors='ignore')}")
        items = data.get("items", [])
        if not items:
            break
        results.extend(items)
        if len(items) < 100:
            break
        start += 100
    return results


def parse_pubdate(s):
    return datetime.strptime(s, "%a, %d %b %Y %H:%M:%S %z")


def filter_by_month(items, year, month):
    start = datetime(year, month, 1, tzinfo=KST)
    end = datetime(year + 1, 1, 1, tzinfo=KST) if month == 12 else datetime(year, month + 1, 1, tzinfo=KST)
    out = []
    for it in items:
        try:
            dt = parse_pubdate(it["pubDate"])
        except Exception:
            continue
        if start <= dt < end:
            it["_dt"] = dt
            out.append(it)
    return out


def dedupe_by_url(items):
    seen = set()
    out = []
    for it in items:
        url = it.get("originallink") or it.get("link")
        if url in seen:
            continue
        seen.add(url)
        out.append(it)
    return out


def fetch_month_articles(keyword, year, month, client_id, client_secret):
    """네이버에서 수집 -> 해당 월 필터 -> URL 중복제거 -> 매체 계산 -> 날짜우선/매체순 정렬"""
    raw = fetch_all_naver(keyword, client_id, client_secret)
    filtered = filter_by_month(raw, year, month)
    filtered = dedupe_by_url(filtered)
    for it in filtered:
        it["_media"] = media_from_link(it.get("originallink") or it.get("link"))
    filtered.sort(key=lambda x: (x["_dt"].date(), media_sort_key(x["_media"]), x["_dt"]))
    return filtered


# ---------- 엑셀 서식 복제 관련 ----------

def duplicate_sheet_safely(wb, src_ws, new_title, insert_index):
    """wb.copy_worksheet()는 하이퍼링크가 있는 셀의 값을 오염시키는 버그가 있어
    직접 셀 단위로 값/서식만 복제하고 하이퍼링크는 의도적으로 가져오지 않음."""
    ws = wb.create_sheet(new_title, insert_index)

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[row_idx].height = dim.height
    for merged_range in list(src_ws.merged_cells.ranges):
        ws.merge_cells(str(merged_range))

    return ws


def find_row_by_prefix(ws, prefix, col=2, search_range=None):
    rows = range(1, ws.max_row + 1) if search_range is None else search_range
    for r in rows:
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip().startswith(prefix):
            return r
    return None


def clear_row_values(ws, row, min_col=2, max_col=11):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None


def build_month_tab(wb, template_sheet, new_month, year, keyword, articles):
    """make_month_tab.py main()의 시트 조립 로직. articles는 fetch_month_articles()로 이미
    수집/정렬된 리스트 (빈 리스트면 --skip-fetch와 동일하게 3번 섹션만 비워둔 채 생성)."""
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"템플릿 시트 '{template_sheet}' 를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")

    new_sheet_name = f"{new_month}월"
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    src_ws = wb[template_sheet]
    template_index = wb.sheetnames.index(template_sheet)
    ws = duplicate_sheet_safely(wb, src_ws, new_sheet_name, template_index)

    ws["B2"] = f"{keyword} {new_month}월 PR리포트"

    # 1. 기사형식별 보도 수: 값만 0으로 초기화
    r1 = find_row_by_prefix(ws, "1.")
    if r1:
        for r in range(r1 + 2, r1 + 8):
            ws.cell(row=r, column=3).value = 0
        ws.cell(row=r1 + 8, column=3).value = 0

    # 2. 보도자료별 게재 현황: 기존 데이터 행 비우기
    r2 = find_row_by_prefix(ws, "2.")
    r3 = find_row_by_prefix(ws, "3.")
    if r2 and r3:
        for r in range(r2 + 2, r3 - 1):
            clear_row_values(ws, r)

    # 3. 주요 기사 리스트: 데이터 행 전체 삭제 후 새 기사로 재생성
    r3 = find_row_by_prefix(ws, "3.")
    r4 = find_row_by_prefix(ws, "4.")
    header_row = r3 + 1
    old_data_start = header_row + 1
    old_data_end = r4 - 2
    old_count = max(0, old_data_end - old_data_start + 1)

    new_count = len(articles)

    saved_style = None
    saved_row_height = None
    if old_count > 0:
        saved_style = {}
        for col in range(2, 12):
            c = ws.cell(row=old_data_start, column=col)
            saved_style[col] = {
                "font": copy.copy(c.font),
                "border": copy.copy(c.border),
                "fill": copy.copy(c.fill),
                "alignment": copy.copy(c.alignment),
                "number_format": c.number_format,
            }
        saved_row_height = ws.row_dimensions[old_data_start].height

    delta = new_count - old_count
    saved_merges = []
    for mc in list(ws.merged_cells.ranges):
        saved_merges.append((mc.min_row, mc.min_col, mc.max_row, mc.max_col))
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    if old_count > 0:
        ws.delete_rows(old_data_start, old_count)
    if new_count > 0:
        ws.insert_rows(old_data_start, new_count)

    cutoff = old_data_start + old_count
    for (mr1, c1, mr2, c2) in saved_merges:
        if mr1 >= cutoff:
            mr1 += delta
            mr2 += delta
        elif mr2 >= old_data_start:
            continue
        ws.merge_cells(start_row=mr1, start_column=c1, end_row=mr2, end_column=c2)

    for i in range(new_count):
        row = old_data_start + i
        if saved_style is not None:
            for col in range(2, 12):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                s = saved_style[col]
                cell.font = copy.copy(s["font"])
                cell.border = copy.copy(s["border"])
                cell.fill = copy.copy(s["fill"])
                cell.alignment = copy.copy(s["alignment"])
                cell.number_format = s["number_format"]
            if saved_row_height is not None:
                ws.row_dimensions[row].height = saved_row_height
        it = articles[i]
        ws.cell(row=row, column=2).value = i + 1
        ws.cell(row=row, column=3).value = it.get("_media") or media_from_link(it.get("originallink") or it.get("link"))
        ws.cell(row=row, column=4).value = clean_html(it["title"])
        ws.cell(row=row, column=7).value = it.get("originallink") or it.get("link")
        ws.cell(row=row, column=10).value = it["_dt"].strftime("%Y-%m-%d")
        ws.cell(row=row, column=11).value = ""
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)

    # 4. 단순 언급 주요 기사: 데이터 행 비우기
    r4 = find_row_by_prefix(ws, "4.")
    r5 = find_row_by_prefix(ws, "5.")
    if r4 and r5:
        for r in range(r4 + 2, r5 - 1):
            clear_row_values(ws, r)

    # 5. 기자 미팅 리스트: 예시 데이터 비우기 (안내문구 행 유지)
    r5 = find_row_by_prefix(ws, "5.")
    r6 = find_row_by_prefix(ws, "6.")
    if r5 and r6:
        for r in range(r5 + 2, r6 - 2):
            clear_row_values(ws, r)

    # 6. 익월 PR 목표: 제목의 월 갱신 + 예시 데이터 비우기
    r6 = find_row_by_prefix(ws, "6.")
    if r6:
        next_month = new_month + 1 if new_month < 12 else 1
        ws.cell(row=r6, column=2).value = f"6. {next_month}월 PR 목표"
        for r in range(r6 + 2, ws.max_row + 1):
            clear_row_values(ws, r)

    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].width = 16

    return ws, new_count


# ---------- classify_articles.py 로직 ----------

def fetch_text(url, timeout=8):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; ReportBot/1.0)"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            html = resp.read(500_000).decode("utf-8", errors="ignore")
    except Exception:
        return None

    html = re.sub(r"<script.*?</script>", " ", html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r"<style.*?</style>", " ", html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;|&quot;|&amp;|&lt;|&gt;|&apos;", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_relevant_paragraphs(text, keyword, window=700, max_occurrences=6):
    count = text.count(keyword)
    if count == 0:
        return [], 0
    snippets = []
    start_search = 0
    for _ in range(max_occurrences):
        idx = text.find(keyword, start_search)
        if idx == -1:
            break
        start = max(0, idx - 50)
        end = min(len(text), idx + window)
        snippets.append(text[start:end])
        start_search = idx + len(keyword)
    return snippets, count


def looks_like_press_release(snippet, keyword):
    if keyword not in snippet:
        return False
    kw = re.escape(keyword)
    patterns = [
        rf"{kw}\s*(은|는|이|가)\s*[^.]{{0,150}}(라고|다고)\s*\d{{0,2}}일?\s*(밝혔다|말했다|전했다)",
        rf"{kw}\s*[가-힣]{{0,10}}\s*(대표|CTO|CEO|본부장)[^.]{{0,20}}(는|은)?[^.]{{0,80}}(말했다|밝혔다|전했다)",
    ]
    return any(re.search(p, snippet) for p in patterns)


def classify_row(title, url, keyword):
    if title.strip().startswith(keyword):
        return "보도자료", "제목이 키워드로 시작"

    text = fetch_text(url)
    if text is None:
        return None, "본문 조회 실패 - 공란 유지"

    snippets, count = extract_relevant_paragraphs(text, keyword)
    if count == 0:
        return None, "본문에서 키워드를 찾지 못함 - 공란 유지"

    for snippet in snippets:
        if looks_like_press_release(snippet, keyword):
            return "보도자료", f"본문에 보도자료 문장 패턴 확인됨 (언급 {count}회)"

    return None, f"보도자료 패턴 미확인 (언급 {count}회) - 공란 유지"


def classify_report(wb, sheet_name, keyword, progress_callback=None):
    """3번 섹션의 K열(기사형식)을 자동 분류. 이미 값 있는 행은 건드리지 않음.
    progress_callback(done, total, title, result)를 매 행마다 호출."""
    ws = wb[sheet_name]

    r3 = find_row_by_prefix(ws, "3.")
    r4 = find_row_by_prefix(ws, "4.")
    if not r3 or not r4:
        raise ValueError("3번/4번 섹션을 찾을 수 없습니다.")

    header_row = r3 + 1
    data_start = header_row + 1
    data_end = r4 - 2

    rows_to_process = []
    for row in range(data_start, data_end + 1):
        title = ws.cell(row=row, column=4).value
        url = ws.cell(row=row, column=7).value
        if not title:
            continue
        current = ws.cell(row=row, column=11).value
        if current:
            continue
        rows_to_process.append((row, title, url))

    filled, blanked = 0, 0
    total = len(rows_to_process)
    for i, (row, title, url) in enumerate(rows_to_process, start=1):
        result, reason = classify_row(title, url, keyword)
        if result:
            ws.cell(row=row, column=11).value = result
            filled += 1
        else:
            blanked += 1
        if progress_callback:
            progress_callback(i, total, title, result or "공란")
        time.sleep(0.5)  # 과도한 요청 방지

    return filled, blanked
