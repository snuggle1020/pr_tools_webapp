# -*- coding: utf-8 -*-
"""
"네이버 기사 수집" 탭 전용 시트 빌더.
지정한 날짜에, 입력한 키워드(여러 개 가능) 각각으로 검색된 네이버 기사를
하나의 표(No/키워드/매체명/제목/URL/게재일자)로 모아 엑셀에 담는다.
"""
from openpyxl.styles import Font, PatternFill, Alignment

from media_order import media_sort_key
from report_sheet import load_domain_map, resolve_media_name, domain_of

HEADER_FILL = PatternFill(fgColor="C5D9F1", fill_type="solid")
FONT_NAME = "맑은 고딕"


def build_daily_sheet(wb, sheet_name, keyword_coverage):
    """keyword_coverage: [(keyword, coverage_list), ...] — 각 키워드별 naver_news.collect_coverage() 결과.
    (검색은 키워드별로 하지만, 시트에는 키워드 열 없이 결과만 합쳐서 보여줌)
    반환값: (unmapped_domains, auto_detected) — report_sheet.build_sheet()와 동일한 의미."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    widths = {"A": 6, "B": 16, "C": 60, "D": 45, "E": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = ["No", "매체명", "제목", "URL", "게재일자"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=1 + i, value=h)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    domain_map = load_domain_map()
    unmapped_domains = set()
    auto_detected = {}

    rows = []
    for keyword, coverage in keyword_coverage:
        for item in coverage:
            domain = domain_of(item["URL"])
            media_name, was_auto = resolve_media_name(item["URL"], domain_map)
            if media_name is None:
                unmapped_domains.add(domain)
                media_name = f"[확인필요:{domain}]"
            elif was_auto:
                auto_detected[domain] = media_name
            rows.append((media_name, item))

    # 날짜순 정렬, 같은 시각이면 매체 우선순위로
    rows.sort(key=lambda r: (r[1]["게재일자"], media_sort_key(r[0])))

    for i, (media_name, item) in enumerate(rows):
        row = 2 + i
        values = [i + 1, media_name, item["제목"], item["URL"], item["게재일자"]]
        for j, v in enumerate(values):
            col = 1 + j
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 3:  # 제목 왼쪽 정렬
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col == 4:  # URL 왼쪽 정렬 + 하이퍼링크
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.hyperlink = v
            if col == 5:  # 게재일자 포맷
                cell.number_format = "yyyy-mm-dd"

    return unmapped_domains, auto_detected
