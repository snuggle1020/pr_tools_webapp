# -*- coding: utf-8 -*-
"""
PR 리포트 도구 웹앱.
- 탭1: 결과보고서 (보도자료 1건 배포 결과 - 네이버 뉴스 수집 후 엑셀 시트 생성)
- 탭2: 월간 PR리포트 (기존 워크북에 새 달 시트 추가 + 기사형식 자동분류)

네이버 API 키는 Streamlit Cloud의 App settings > Secrets 에
NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 로 등록해두면 자동으로 읽힘.
"""
import io
import os
from datetime import date

import openpyxl
import streamlit as st

from naver_news import collect_coverage
from report_sheet import build_sheet, derive_company, derive_keyword
from pr_report_core import fetch_month_articles, build_month_tab, classify_report
from daily_collect_sheet import build_daily_sheet

REPORT_WINDOW_DAYS = 3  # 배포일 이후 며칠까지 검색할지 (고정)

st.set_page_config(page_title="오픈피알_PR 보고서", page_icon="📰", layout="wide")


def _load_naver_credentials():
    cid = os.environ.get("NAVER_CLIENT_ID")
    csecret = os.environ.get("NAVER_CLIENT_SECRET")
    if not cid or not csecret:
        try:
            cid = st.secrets["NAVER_CLIENT_ID"]
            csecret = st.secrets["NAVER_CLIENT_SECRET"]
            os.environ["NAVER_CLIENT_ID"] = cid
            os.environ["NAVER_CLIENT_SECRET"] = csecret
        except Exception:
            cid, csecret = None, None
    return cid, csecret


NAVER_CLIENT_ID, NAVER_CLIENT_SECRET = _load_naver_credentials()

st.title("📰 PR 보고서 생성기")

if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
    st.error(
        "네이버 API 키가 설정되어 있지 않습니다. 관리자에게 문의하거나, "
        "로컬 실행 시 `.streamlit/secrets.toml`에 NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 을 등록하세요."
    )

tab1, tab2, tab3 = st.tabs(["📄 보도자료 결과보고서", "📅 월간 PR리포트", "🔎 네이버 기사 수집"])

# ------------------------------------------------------------------
# 탭 1: 결과보고서
# ------------------------------------------------------------------
with tab1:
    st.caption(
        "보도자료 제목만 넣으면 기업명/검색어는 자동으로 뽑아냅니다. "
        f"배포일 당일부터 {REPORT_WINDOW_DAYS}일 뒤까지 게재된 기사만 모읍니다."
    )

    with st.form("report_form"):
        title = st.text_input("보도자료 제목", placeholder="예: 딥파인, 국방·산업 현장용 스마트글래스 기반 AI 에이전트 개발 나선다")
        distribution_date = st.date_input("배포일", value=date.today())
        existing_file = st.file_uploader(
            "기존 결과보고서 파일 (파일을 추가하면 맨 왼쪽에 새 시트를 추가합니다 / 없으면 새 파일을 생성합니다)",
            type=["xlsx"],
            key="report_existing_wb",
        )
        submitted = st.form_submit_button("리포트 생성", type="primary")

    if submitted:
        if not title:
            st.error("보도자료 제목을 입력하세요.")
        elif not NAVER_CLIENT_ID:
            st.error("네이버 API 키가 없어 진행할 수 없습니다.")
        else:
            company = derive_company(title)
            keyword = derive_keyword(title)
            st.caption(f"자동 인식: 기업명 **{company}** / 검색어 **{keyword}**")

            dist_str = distribution_date.strftime("%Y-%m-%d")
            with st.spinner(f"네이버에서 '{keyword}' 검색 중..."):
                try:
                    coverage = collect_coverage(
                        keyword, dist_str, window_days=REPORT_WINDOW_DAYS, require_text=company
                    )
                except Exception as e:
                    st.error(f"뉴스 수집 실패: {e}")
                    coverage = None

            if coverage is not None:
                sheet_name = distribution_date.strftime("%y%m%d")
                if existing_file is not None:
                    wb = openpyxl.load_workbook(existing_file)
                else:
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)

                with st.spinner("고정 목록에 없는 매체는 기사 페이지에서 매체명 자동 인식 중..."):
                    unmapped, auto_detected = build_sheet(wb, sheet_name, title, company, dist_str, coverage)
                wb.active = 0  # 새로 추가된(맨 왼쪽) 시트가 열리도록

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                request_date_str = date.today().strftime("%Y%m%d")
                filename = f"결과보고서_{title}_{request_date_str}.xlsx"

                st.success(f"완료! 기사 {len(coverage)}건 수집됨 (시트: {sheet_name}).")
                if auto_detected:
                    lines = "\n".join(f"- {d} → {n}" for d, n in sorted(auto_detected.items()))
                    st.info(
                        "다음 매체는 고정 목록에 없어서 기사 페이지에서 자동으로 인식했습니다 "
                        "(맞는지 확인해보고, domain_to_media.json에 추가해두면 다음부턴 더 빨라져요):\n"
                        + lines
                    )
                if unmapped:
                    st.warning(
                        "매체명을 자동으로도 못 찾은 도메인: " + ", ".join(sorted(unmapped))
                        + " — 직접 확인해서 domain_to_media.json에 추가해주세요."
                    )
                st.download_button(
                    "⬇️ 엑셀 다운로드",
                    data=buf,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

# ------------------------------------------------------------------
# 탭 2: 월간 PR리포트
# ------------------------------------------------------------------
with tab2:
    st.caption("기존 파일에 새 달 시트를 추가하고, 기사형식이 보도자료인 것만 자동분류합니다.")

    template_file = st.file_uploader(
        "기존 PR리포트 파일 업로드", type=["xlsx"], key="pr_template"
    )

    sheet_options = []
    if template_file is not None:
        try:
            template_file.seek(0)
            _preview_wb = openpyxl.load_workbook(template_file, read_only=True)
            sheet_options = [s for s in _preview_wb.sheetnames if s.endswith("월")]
            _preview_wb.close()
        except Exception as e:
            st.error(f"워크북을 열 수 없습니다: {e}")

    keyword2 = st.text_input("키워드", value="딥파인", key="pr_keyword")
    c1, c2 = st.columns(2)
    year = c1.number_input("연도", min_value=2000, max_value=2100, value=date.today().year, step=1, key="pr_year")
    month = c2.number_input("새로 만들 달", min_value=1, max_value=12, value=date.today().month, step=1, key="pr_month")

    if sheet_options:
        template_sheet_name = st.selectbox(
            "서식을 복제할 기존 월 시트 (가장 최근 달이 기본 선택됨)", sheet_options, index=0
        )
    else:
        template_sheet_name = st.text_input("서식을 복제할 기존 월 시트 이름 (예: 6월)", key="pr_sheet_manual")

    do_classify = st.checkbox("기사형식 자동 분류도 실행", value=True, key="pr_classify")

    if st.button("PR리포트 생성", type="primary", key="pr_submit"):
        if template_file is None:
            st.error("기존 워크북을 업로드하세요.")
        elif not template_sheet_name:
            st.error("서식을 복제할 월 시트 이름을 지정하세요.")
        elif not keyword2:
            st.error("키워드를 입력하세요.")
        elif not NAVER_CLIENT_ID:
            st.error("네이버 API 키가 없어 진행할 수 없습니다.")
        else:
            template_file.seek(0)
            wb = openpyxl.load_workbook(template_file)

            if template_sheet_name not in wb.sheetnames:
                st.error(f"'{template_sheet_name}' 시트를 찾을 수 없습니다. 현재 시트: {wb.sheetnames}")
            else:
                with st.spinner(f"네이버에서 '{keyword2}' 검색 중..."):
                    try:
                        articles = fetch_month_articles(
                            keyword2, int(year), int(month), NAVER_CLIENT_ID, NAVER_CLIENT_SECRET
                        )
                    except Exception as e:
                        st.error(f"뉴스 수집 실패: {e}")
                        articles = None

                if articles is not None:
                    ws, new_count = build_month_tab(wb, template_sheet_name, int(month), int(year), keyword2, articles)
                    sheet_name = f"{int(month)}월"
                    st.success(f"'{sheet_name}' 시트 생성 완료. 3번 섹션에 기사 {new_count}건 채움.")

                    if do_classify and new_count > 0:
                        progress_bar = st.progress(0.0, text="기사형식 분류 준비 중...")

                        def _cb(done, total, t, result):
                            frac = done / total if total else 1.0
                            progress_bar.progress(frac, text=f"[{done}/{total}] {result} - {t[:30]}")

                        filled, blanked = classify_report(wb, sheet_name, keyword2, progress_callback=_cb)
                        progress_bar.progress(1.0, text="분류 완료")
                        st.info(
                            f"확정 보도자료로 채운 행: {filled}건 / 공란으로 남은 행: {blanked}건 "
                            "(공란은 인터뷰/기고/기획(오픈피알)/기획(외부)/단순언급 중 직접 확인해서 채워주세요)"
                        )

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    filename = f"{int(year)}년_{int(month)}월_{keyword2}_PR리포트.xlsx"

                    st.download_button(
                        "⬇️ 엑셀 다운로드",
                        data=buf,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    st.caption(
                        "1/2/4/5/6번 섹션은 지난달 데이터가 지워진 빈 템플릿 상태이니 직접 채워주세요. "
                        "다음 달 작업을 위해 이번에 만든 파일을 템플릿으로 보관해두세요."
                    )

# ------------------------------------------------------------------
# 탭 3: 네이버 기사 수집
# ------------------------------------------------------------------
with tab3:
    st.caption("지정한 기간에 게재된, 키워드가 포함된 네이버 기사를 모두 모아 엑셀로 만듭니다. 리포트 양식 없이 목록만 뽑습니다.")

    with st.form("daily_form"):
        date_range = st.date_input(
            "기간",
            value=(date.today(), date.today()),
            key="daily_date_range",
        )
        keywords_text = st.text_area(
            "키워드 (여러 개면 줄바꿈으로 구분)",
            placeholder="예)딥파인\n엠아이큐브솔루션",
            key="daily_keywords",
        )
        submitted3 = st.form_submit_button("기사 수집", type="primary")

    if submitted3:
        keywords = [k.strip() for k in keywords_text.splitlines() if k.strip()]
        if isinstance(date_range, (list, tuple)):
            start_date = date_range[0]
            end_date = date_range[1] if len(date_range) > 1 else date_range[0]
        else:
            start_date = end_date = date_range

        if not keywords:
            st.error("키워드를 최소 1개 입력하세요.")
        elif start_date > end_date:
            st.error("시작일이 종료일보다 늦을 수 없습니다.")
        elif not NAVER_CLIENT_ID:
            st.error("네이버 API 키가 없어 진행할 수 없습니다.")
        else:
            start_str = start_date.strftime("%Y-%m-%d")
            window_days = (end_date - start_date).days
            keyword_coverage = []
            with st.spinner(f"네이버에서 {len(keywords)}개 키워드 검색 중..."):
                for kw in keywords:
                    try:
                        cov = collect_coverage(kw, start_str, window_days=window_days, require_text=kw)
                    except Exception as e:
                        st.error(f"'{kw}' 검색 실패: {e}")
                        cov = []
                    keyword_coverage.append((kw, cov))

            total = sum(len(cov) for _, cov in keyword_coverage)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            if start_date == end_date:
                sheet_name = start_date.strftime("%y%m%d")
                date_str = start_date.strftime("%Y%m%d")
            else:
                sheet_name = f"{start_date.strftime('%y%m%d')}-{end_date.strftime('%y%m%d')}"
                date_str = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"

            with st.spinner("고정 목록에 없는 매체는 기사 페이지에서 매체명 자동 인식 중..."):
                unmapped, auto_detected = build_daily_sheet(wb, sheet_name, keyword_coverage)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"네이버기사수집_{date_str}.xlsx"

            st.success(f"완료! 총 {total}건 수집됨.")
            for kw, cov in keyword_coverage:
                st.caption(f"- {kw}: {len(cov)}건")
            if auto_detected:
                lines = "\n".join(f"- {d} → {n}" for d, n in sorted(auto_detected.items()))
                st.info(
                    "다음 매체는 고정 목록에 없어서 기사 페이지에서 자동으로 인식했습니다 "
                    "(맞는지 확인해보고, domain_to_media.json에 추가해두면 다음부턴 더 빨라져요):\n"
                    + lines
                )
            if unmapped:
                st.warning(
                    "매체명을 자동으로도 못 찾은 도메인: " + ", ".join(sorted(unmapped))
                    + " — 직접 확인해서 domain_to_media.json에 추가해주세요."
                )
            st.download_button(
                "⬇️ 엑셀 다운로드",
                data=buf,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="daily_download",
            )
