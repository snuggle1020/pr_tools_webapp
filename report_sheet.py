# -*- coding: utf-8 -*-
"""
결과보고서(보도자료 배포 결과) 시트 빌더.
결과보고서/generate_report.py 의 build_sheet() 로직을 그대로 옮김.

매체명 조회 순서: domain_to_media.json 고정 목록 -> (없으면) 기사 페이지의
og:site_name 메타태그를 직접 읽어서 자동 인식 -> 그래도 실패하면 [확인필요:도메인].
"""
import json
import re
from pathlib import Path
from urllib.parse import urlparse
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment

from media_order import media_sort_key
from pr_report_core import _fetch_site_name

HERE = Path(__file__).parent
DOMAIN_MAP_PATH = HERE / "domain_to_media.json"

HEADER_FILL = PatternFill(fgColor="C5D9F1", fill_type="solid")
FONT_NAME = "맑은 고딕"

# 이번 실행 중 새로 자동 인식한 도메인 캐시 (같은 도메인 반복 조회 방지용)
_auto_detected_cache = {}


def load_domain_map():
    if DOMAIN_MAP_PATH.exists():
        with open(DOMAIN_MAP_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def domain_of(url: str) -> str:
    return urlparse(url).netloc.replace("www.", "")


def derive_company(title: str) -> str:
    """보도자료 제목은 관례적으로 '{기업명}, ...' 형식이라 쉼표 앞부분을 기업명으로 사용."""
    for comma in [",", "，"]:
        if comma in title:
            return title.split(comma, 1)[0].strip()
    return title.strip()


def derive_keyword(title: str) -> str:
    """네이버 뉴스 검색어. 실측 결과 제목 전체를 검색어로 쓰면 단어가 많아져
    오히려 검색 결과가 급감함(12건 -> 3건). 기업명 하나만 검색어로 쓰는 쪽이
    재현율이 훨씬 높고, 관련성 없는 기사는 collect_coverage의 require_text=기업명
    필터가 걸러줌."""
    return derive_company(title)


def resolve_media_name(url: str, domain_map: dict):
    """(매체명, 자동인식여부) 반환. 고정 목록에 없으면 페이지에서 직접 읽어서 자동 인식 시도."""
    domain = domain_of(url)
    if domain in domain_map:
        return domain_map[domain], False  # 고정 목록에서 바로 찾음
    if domain in _auto_detected_cache:
        return _auto_detected_cache[domain], True
    name = _fetch_site_name(url)
    if name:
        _auto_detected_cache[domain] = name
        return name, True
    return None, True  # 자동 인식도 실패


def build_sheet(wb, sheet_name, title, company, distribution_date, coverage):
    """coverage: naver_news.collect_coverage()가 반환한 리스트. 반환값: 매체명 미확인 도메인 집합."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)  # 맨 왼쪽에 생성

    ws.sheet_view.zoomScale = 85  # 파일을 열었을 때 기본 줌 85%

    widths = {"A": 8.8, "B": 9.0, "C": 17.1, "D": 56.6, "E": 50.2, "F": 14.3, "G": 20.4}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 제목 행 (하늘색 음영 B~G까지)
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws[f"{col}1"].fill = HEADER_FILL
    ws["B1"] = "보도자료 배포 결과 보고서"
    ws["B1"].font = Font(name=FONT_NAME, size=15, bold=True)
    ws["B1"].alignment = Alignment(horizontal=None, vertical="center")
    ws.row_dimensions[1].height = 28.2

    meta_rows = [
        (3, "제목", title),
        (4, "기업명", company),
        (5, "배포일", datetime.strptime(distribution_date, "%Y-%m-%d")),
    ]
    for row, label, value in meta_rows:
        b = ws.cell(row=row, column=2, value=label)
        b.font = Font(name=FONT_NAME, size=10)
        b.fill = HEADER_FILL
        b.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=row, column=3, value=value)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if row == 5:
            c.number_format = "yyyy-mm-dd"

    headers = ["No", "매체명", "제목", "URL", "게재 일자", "게재 포털"]
    for i, h in enumerate(headers):
        col = 2 + i
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    domain_map = load_domain_map()
    unmapped_domains = set()
    auto_detected = {}  # domain -> media_name (고정 목록엔 없지만 자동 인식 성공)

    enriched = []
    for item in coverage:
        domain = domain_of(item["URL"])
        media_name, was_auto = resolve_media_name(item["URL"], domain_map)
        if media_name is None:
            unmapped_domains.add(domain)
            media_name = f"[확인필요:{domain}]"
        elif was_auto:
            auto_detected[domain] = media_name
        enriched.append((media_name, item))

    # 매체 우선순위로 정렬(통신사>일간지>경제지>IT전문지>산업지>온라인매체), 같은 매체면 날짜순
    enriched.sort(
        key=lambda pair: (media_sort_key(pair[0]), pair[1]["게재일자"])
    )

    for i, (media_name, item) in enumerate(enriched):
        row = 8 + i
        values = [
            i + 1,
            media_name,
            item["제목"],
            item["URL"],
            item["게재일자"],
            item["게재포털"],
        ]
        for j, v in enumerate(values):
            col = 2 + j
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 4:  # 제목 열(D) 왼쪽 정렬
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col == 5:  # URL 열(E) 왼쪽 정렬 + 하이퍼링크
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.hyperlink = v
            if col == 6:  # 게재일자 열(F) 포맷: 2026-07-02
                cell.number_format = "yyyy-mm-dd"

    return unmapped_domains, auto_detected
